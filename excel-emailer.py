""" program that reads an excel file of recipients and sends them an email via a scheduler function. 
Be sure to configure the scheduler() and pass_criteria() functions for your base case.
Also be sure to check that the excel sheet is configured the same way as the example.  Changing cell locations will break it. 
Currently configured for Gmail, but comments show how to use Outlook on local server as well. 
"""

import openpyxl # dl pip3 install openpyxl==2.6.2 so that code in Automate the Boring Stuff works
from random import choice
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime
import schedule
import time
import imghdr #used to get image file type automatically, for email attachments

# only runs main() at the specified time
def scheduler():
    schedule.every().day.at("17:13:45").do(main)
    while True:
        schedule.run_pending()
        time.sleep(1)

# only send emails to those people who meet certain criteria
def pass_criteria(cat): 
    return cat == "Signage"

# open the excel customer database file and read it into memory
def read_excel():
    wb = openpyxl.load_workbook('testexcelwb.xlsx')
    sheet1 = wb['Customer_DB']
    sheet2 = wb['email_opts']
    html_options = sheet2['B2'].value, sheet2['B3'].value
    #print(html_options)

    recipients = []
    companies = []
    for row in range(2, sheet1.max_row + 1): #sheet.max_row
        cat = sheet1['F' + str(row)].value
        if pass_criteria(cat):
            fn = sheet1['A' + str(row)].value
            ln = sheet1['B' + str(row)].value
            company = sheet1['C' + str(row)].value
            email = sheet1['D' + str(row)].value

            company_info = {}
            company_info['company'] = company
            company_info['first_name'] = fn
            company_info['last_name'] = ln
            company_info['email'] = email
            company_info['index'] = row
            # choose alternative email text for customers at the same company (appears more unique/authentic)
            if company not in companies:
                companies.append(company)
                company_info['html'] = html_options[0]
            else:
                company_info['html'] = html_options[1]
            recipients.append(company_info)

    return recipients

# write the date/time next to each customer that received an email, to keep track of progress
def write_excel(i, e = None):
    print("Writing Excel...")        
    wb = openpyxl.load_workbook('testexcelwb.xlsx')
    sheet = wb['Customer_DB']
    if e != None:
        sheet['E' + str(i)] = str(e)
    else:
        sheet['E' + str(i)] = datetime.now()
    wb.save('testexcelwb.xlsx')
    print("Excel Updated!")

# send emails
def send_email(recipients):
    # Outlook: from_email = "fake@email.com" hardcode it in if using a local server
    from_email = os.environ.get('EMAIL_USER')
    # Outlook: comment out password and port lines
    password = os.environ.get('EMAIL_PASS')
    port = 587
    # Outlook: change server to 'smtp.company.local' where company is actual company name
    server = 'smtp.gmail.com'

    # prep a pdf attachment
    with open('DP8405_TDS.pdf', 'rb') as f:
        file_data = f.read()
        file_type = imghdr.what(f.name)
        file_name = f.name
    attachedfile = MIMEApplication(file_data, _subtype = "pdf")
    attachedfile.add_header('content-disposition', 'attachment', filename = file_name)
    
    # loop through all email recipients in the dict list and send personalized emails to them
    for recipient in recipients:
        to_email = recipient['email']
        first_name = recipient['first_name']
        last_name = recipient['last_name']
        company = recipient['company']
        index = recipient['index']
        html = recipient['html']

        # define and setup email message. Must be inside the for loop
        message = MIMEMultipart("alternative")
        message["From"] = from_email
        message["To"] = to_email
        message["Subject"] = f"Adhesive Alternatives for {company}"
        html = html.format(last_name=last_name, first_name=first_name, company=company, to_email=to_email)
        text = html.format(last_name=last_name, first_name=first_name, company=company, to_email=to_email)
        message.attach(attachedfile)

        # Turn these into plain/html MIMEText objects
        part1 = MIMEText(text, "plain")
        part2 = MIMEText(html, "html")

        # Add HTML/plain-text parts to MIMEMultipart message
        # The email client will try to render the last part first
        message.attach(part1)
        message.attach(part2)
        print(f"Preparing to send email to {first_name}")
        try:
            # Outlook: remove port from the following line.  It defaults to the correct one for local servers (25 or 24)
            smtpObj = smtplib.SMTP(server, port)
            smtpObj.ehlo()
            smtpObj.starttls()
            # Outlook: comment out the login line. Not needed on local network.
            smtpObj.login(from_email, password)
            print("Sending Email...")
            smtpObj.sendmail(from_email, to_email, message.as_string()) #must send msg as.string() when using HTML and plaintext options
            print('Email Sent!')
            write_excel(index)
        except Exception as e:
            print(e)
            write_excel(index, e)
        finally:
            smtpObj.quit()

def main():
    recipients = read_excel()
    send_email(recipients)

if __name__ == '__main__':
    scheduler()
    # main()